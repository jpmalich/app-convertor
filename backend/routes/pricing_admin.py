"""Supplier-admin pricing-update tools.

Three workflows:
  1. **% Bump**     — apply a percentage change across tiers/sections (preview + apply).
  2. **CSV/XLSX**   — upload a price sheet; backend computes a diff against current tiers.
  3. **Export**     — download current prices as CSV (edit in Excel + re-upload).

All endpoints are token-gated via `check_admin_token` (same as the rest of /admin/*).

Pattern: the preview endpoints return a `changes` array with `old` and `new` values
per cell. The frontend renders that as a diff table; the user then POSTs the SAME
changes array back to `/admin/pricing/apply` to commit. This makes the apply step a
trivial bulk update and lets the frontend hold the staged diff in component state.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from db import db
from deps import check_admin_token
from services import ensure_tiers_seeded

router = APIRouter()

# Accepted columns when parsing uploads. We're permissive about header case
# (the supplier may export from Excel with mixed casing) but strict about the
# fields themselves so a malformed file fails fast with a clear error.
CSV_COLUMNS = ["tier", "section", "name", "unit", "mat", "lab"]


# ---------------------------------------------------------------------------
# Pydantic
# ---------------------------------------------------------------------------
class BumpScope(BaseModel):
    tier_ids: Optional[list[str]] = None       # default: all tiers
    section_titles: Optional[list[str]] = None  # default: all sections


class BumpIn(BaseModel):
    percent: float                          # e.g. 4.5 means +4.5%; -2 means -2%
    scope: Optional[BumpScope] = None
    target: str = "mat"                     # "mat" | "lab" | "both"


class PriceChange(BaseModel):
    """A single proposed price change. Stored in component state between
    preview + apply so the apply step doesn't need to recompute anything."""
    tier_id: str
    tier_name: str
    section: str
    name: str
    unit: str
    field: str          # "mat" | "lab"
    old: float
    new: float


class ApplyIn(BaseModel):
    changes: list[PriceChange]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _round2(x: float) -> float:
    """Round to cents. Matches how price sheets are typically displayed."""
    return round(float(x) + 0.0, 2)


async def _load_all_tiers() -> list[dict]:
    await ensure_tiers_seeded()
    cursor = db.price_tiers.find({}, {"_id": 0}).sort("name", 1)
    return await cursor.to_list(50)


def _build_bump_changes(tiers: list[dict], body: BumpIn) -> list[dict]:
    multiplier = 1 + (body.percent / 100.0)
    scope = body.scope or BumpScope()
    targets = ("mat", "lab") if body.target == "both" else (body.target,)
    if body.target not in {"mat", "lab", "both"}:
        raise HTTPException(status_code=400, detail="target must be 'mat', 'lab', or 'both'")

    tier_filter = set(scope.tier_ids or [])
    section_filter = set(scope.section_titles or [])

    out: list[dict] = []
    for tier in tiers:
        if tier_filter and tier["id"] not in tier_filter:
            continue
        for sec in tier.get("sections", []) or []:
            if section_filter and sec["title"] not in section_filter:
                continue
            for it in sec.get("items", []) or []:
                for field in targets:
                    old = float(it.get(field, 0) or 0)
                    new = _round2(old * multiplier)
                    if abs(new - old) < 0.005:
                        continue  # no visible change at the cent
                    out.append({
                        "tier_id": tier["id"],
                        "tier_name": tier["name"],
                        "section": sec["title"],
                        "name": it["name"],
                        "unit": it.get("unit", ""),
                        "field": field,
                        "old": old,
                        "new": new,
                    })
    return out


def _parse_upload(filename: str, raw: bytes) -> tuple[list[dict], list[dict]]:
    """Returns (rows, unmatched_columns_errors). Rows are normalized dicts with
    keys matching CSV_COLUMNS. We support .csv and .xlsx."""
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    rows: list[dict] = []

    if ext in {"xlsx", "xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise HTTPException(status_code=500, detail="openpyxl not installed") from e
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if header_row is None:
                header_row = [str(c or "").strip().lower() for c in row]
                missing = [c for c in CSV_COLUMNS if c not in header_row]
                if missing:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Spreadsheet missing required columns: {', '.join(missing)}",
                    )
                continue
            d = {header_row[i]: row[i] for i in range(min(len(header_row), len(row)))}
            if not any(d.get(c) for c in ("tier", "section", "name")):
                continue
            rows.append(d)
    else:
        # Treat anything else as CSV — works for .csv and .txt sheets.
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="Empty CSV (no header row)")
        normalized = [(f or "").strip().lower() for f in reader.fieldnames]
        missing = [c for c in CSV_COLUMNS if c not in normalized]
        if missing:
            raise HTTPException(
                status_code=400,
                detail=f"CSV missing required columns: {', '.join(missing)}",
            )
        # Rebuild reader with normalized headers
        text2 = ",".join(normalized) + "\n" + text.split("\n", 1)[1]
        reader = csv.DictReader(io.StringIO(text2))
        for r in reader:
            rows.append(r)
    return rows, []


def _diff_upload(tiers: list[dict], rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Compute a diff between uploaded rows and the current tiers in DB.
    Returns (changes, unmatched). Unmatched rows are items where the (tier, section, name)
    triple didn't resolve to a current catalog item — surfaced in the UI so the
    user can fix typos before applying."""
    # Build a lookup: (tier_name_lower, section_lower, name_lower) -> (tier_id, item)
    lookup: dict[tuple[str, str, str], tuple[str, str, dict, str]] = {}
    for tier in tiers:
        for sec in tier.get("sections", []) or []:
            for it in sec.get("items", []) or []:
                key = (
                    tier["name"].strip().lower(),
                    sec["title"].strip().lower(),
                    it["name"].strip().lower(),
                )
                lookup[key] = (tier["id"], tier["name"], it, sec["title"])

    changes: list[dict] = []
    unmatched: list[dict] = []
    for i, r in enumerate(rows, start=2):  # row 1 was the header
        tier_v = (r.get("tier") or "").strip()
        section_v = (r.get("section") or "").strip()
        name_v = (r.get("name") or "").strip()
        if not (tier_v and section_v and name_v):
            unmatched.append({"row": i, "reason": "Missing tier/section/name"})
            continue
        key = (tier_v.lower(), section_v.lower(), name_v.lower())
        found = lookup.get(key)
        if not found:
            unmatched.append({
                "row": i, "tier": tier_v, "section": section_v, "name": name_v,
                "reason": "No matching catalog item — check spelling/section",
            })
            continue
        tier_id, tier_name, item, section_title = found

        for field in ("mat", "lab"):
            raw_val = r.get(field)
            if raw_val in (None, ""):
                continue
            try:
                new = _round2(float(str(raw_val).replace("$", "").replace(",", "")))
            except (TypeError, ValueError):
                unmatched.append({
                    "row": i, "tier": tier_v, "section": section_v, "name": name_v,
                    "reason": f"Non-numeric {field}: {raw_val!r}",
                })
                continue
            old = float(item.get(field, 0) or 0)
            if abs(new - old) < 0.005:
                continue
            changes.append({
                "tier_id": tier_id,
                "tier_name": tier_name,
                "section": section_title,
                "name": item["name"],
                "unit": item.get("unit", ""),
                "field": field,
                "old": old,
                "new": new,
            })
    return changes, unmatched


async def _apply_changes(changes: list[dict]) -> int:
    """Apply a list of changes to the price_tiers collection.
    Groups by tier_id so each tier is written exactly once."""
    by_tier: dict[str, list[dict]] = {}
    for c in changes:
        by_tier.setdefault(c["tier_id"], []).append(c)

    applied = 0
    for tier_id, tier_changes in by_tier.items():
        tier = await db.price_tiers.find_one({"id": tier_id}, {"_id": 0})
        if not tier:
            continue
        # Index items by (section_title, name) so we can patch in place.
        sections = tier.get("sections", []) or []
        sec_index = {s["title"]: s for s in sections}
        for c in tier_changes:
            sec = sec_index.get(c["section"])
            if not sec:
                continue
            for it in sec.get("items", []) or []:
                if it["name"] == c["name"]:
                    it[c["field"]] = _round2(c["new"])
                    applied += 1
                    break
        await db.price_tiers.update_one(
            {"id": tier_id},
            {"$set": {
                "sections": sections,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }},
        )
    return applied


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.post("/admin/pricing/preview-bump")
async def preview_bump(body: BumpIn, request: Request):
    """Compute a diff for a % bump WITHOUT touching the database."""
    check_admin_token(request)
    tiers = await _load_all_tiers()
    changes = _build_bump_changes(tiers, body)
    return {"changes": changes}


@router.post("/admin/pricing/upload")
async def upload_pricing(request: Request, file: UploadFile = File(...), commit: str = Form("false")):
    """Preview an uploaded CSV/XLSX. Pass `commit=true` to apply in the same call
    (skips the two-step preview-then-apply flow for trusted scripts), otherwise
    returns the diff for the UI to render before the user clicks Apply."""
    check_admin_token(request)
    raw = await file.read()
    if len(raw) > 4 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="File too large (>4MB)")
    rows, _ = _parse_upload(file.filename or "", raw)
    tiers = await _load_all_tiers()
    changes, unmatched = _diff_upload(tiers, rows)
    applied = 0
    if commit.lower() == "true":
        applied = await _apply_changes(changes)
    return {"changes": changes, "unmatched": unmatched, "applied": applied}


@router.post("/admin/pricing/apply")
async def apply_changes(body: ApplyIn, request: Request):
    """Commit a previewed changeset. Frontend takes the array returned by
    /preview-bump or /upload, lets the user review, then POSTs it here."""
    check_admin_token(request)
    if not body.changes:
        return {"applied": 0}
    applied = await _apply_changes([c.model_dump() for c in body.changes])
    return {"applied": applied}


@router.get("/admin/pricing/export")
async def export_pricing(request: Request):
    """Download current prices as CSV. Howard edits in Excel and re-uploads."""
    check_admin_token(request)
    tiers = await _load_all_tiers()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_COLUMNS)
    for tier in tiers:
        for sec in tier.get("sections", []) or []:
            for it in sec.get("items", []) or []:
                writer.writerow([
                    tier["name"], sec["title"], it["name"], it.get("unit", ""),
                    f"{float(it.get('mat', 0) or 0):.2f}",
                    f"{float(it.get('lab', 0) or 0):.2f}",
                ])
    buf.seek(0)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="pricing-{today}.csv"'},
    )
